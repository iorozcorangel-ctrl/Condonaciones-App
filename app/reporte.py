"""
================================================================
  GENERADOR DEL REPORTE EXCEL FINAL
  Produce el archivo con las 63 columnas y formato definido
================================================================
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date as Date
from app.config import COL_BI, COL_TAB, HEADER_COLOR_REPORTE
from app.validaciones import to_date, to_float, to_int, generar_comentario


_THIN   = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_HDR_FILL  = PatternFill("solid", start_color=HEADER_COLOR_REPORTE)
_HDR_ALIGN = Alignment(horizontal="center", vertical="center",
                        wrap_text=True)
_DAT_FONT  = Font(name="Arial", size=10)
_DAT_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_MONEY_FMT = '"$"#,##0.00'
_DATE_FMT  = "DD/MM/YYYY"


def _hdr(ws, row, col, value, colspan=1):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _HDR_FONT
    cell.fill      = _HDR_FILL
    cell.alignment = _HDR_ALIGN
    cell.border    = _BORDER
    if colspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col + colspan - 1
        )
    return cell


def _dat(ws, row, col, value, money=False, fmt_date=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _DAT_FONT
    cell.alignment = _DAT_ALIGN
    cell.border    = _BORDER
    if money:
        cell.number_format = _MONEY_FMT
    if fmt_date and isinstance(value, Date):
        cell.number_format = _DATE_FMT
    return cell


def generar_reporte(df_tab, df_bi, desfases, montos, fecha_solicitud_global,
                    fecha_revision, nc_cliente, fechas_por_contenedor,
                    ruta_salida):
    """
    Genera el reporte Excel final con todas las columnas documentadas.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Condonaciones"

    # ── Fila 1: Encabezados ────────────────────────────────────
    # Las columnas con celdas combinadas se definen como colspan=2
    col = 1
    _hdr(ws, 1, col,  "Contenedor");                                          col += 1
    _hdr(ws, 1, col,  "Procedente/No Procede\n(PRO, NO PRO)");                col += 1
    _hdr(ws, 1, col,  "Fecha de Solicitud\nde Nota de Crédito");              col += 1
    _hdr(ws, 1, col,  "Fecha Revisión NC");                                   col += 1
    _hdr(ws, 1, col,  "Monto Solicitado\npor el Cliente");                    col += 1
    _hdr(ws, 1, col,  "LINER");                                               col += 1
    _hdr(ws, 1, col,  "LINE OP");                                             col += 1
    _hdr(ws, 1, col,  "Fecha de Ingreso a la Terminal", colspan=2);           col += 2
    _hdr(ws, 1, col,  "Fecha de Revalidación\npor A.A. con La Naviera", colspan=2); col += 2
    _hdr(ws, 1, col,  "Fecha de Solicitud\nde Previo por A.A.", colspan=2);   col += 2
    _hdr(ws, 1, col,  "Pasó de los 3 Días\nHábiles Operables (LUN/SÁB)");    col += 1
    _hdr(ws, 1, col,  "Fecha Otorgada por CMSA\n(Fecha de Servicio)", colspan=2); col += 2
    _hdr(ws, 1, col,  "Fecha de Cancelación\n(a petición del cliente)");      col += 1
    _hdr(ws, 1, col,  "¿Cuántos Servicios\nProgramados?");                    col += 1
    _hdr(ws, 1, col,  "Cuántos Días de\nDesfase en Previo?");                 col += 1
    _hdr(ws, 1, col,  "Cuántos Días de Intento\nde Programación");            col += 1
    _hdr(ws, 1, col,  "Fecha de Liberación\npor A.A. (ISCS)", colspan=2);     col += 2
    _hdr(ws, 1, col,  "FFCC");                                                col += 1
    _hdr(ws, 1, col,  "Fecha de Ferromex", colspan=2);                        col += 2
    _hdr(ws, 1, col,  "Pasó de los 3 Días\npara Cargar a FFCC");              col += 1
    _hdr(ws, 1, col,  "Fecha de Carga\na Góndola por CMSA", colspan=2);       col += 2
    _hdr(ws, 1, col,  "Cuántos Días de Desfase\nen Carga a Góndola?");        col += 1
    _hdr(ws, 1, col,  "Fecha de Programación\nde Entrega por A.A.\n(Camión, ISCS)", colspan=2); col += 2
    _hdr(ws, 1, col,  "Fecha de Autorización\npor La Naviera\n(si es Liner)"); col += 1
    _hdr(ws, 1, col,  "Cantidad de\nProgramaciones");                         col += 1
    _hdr(ws, 1, col,  "Fecha Otorgada por CMSA\npara Entrega a Camión", colspan=2); col += 2
    _hdr(ws, 1, col,  "Pasó de los 2 Días Hábiles\nFecha Otorgada CMSA\n(Entrega Camión)"); col += 1
    _hdr(ws, 1, col,  "Fecha Real de Entrega\na Camión (Time Out)", colspan=2); col += 2
    _hdr(ws, 1, col,  "Cuántos Días de\nDesfase en Entrega?");                col += 1
    _hdr(ws, 1, col,  "Comentarios\n(Investigación)");                        col += 1
    _hdr(ws, 1, col,  "Tipo de\nContenedor");                                 col += 1
    _hdr(ws, 1, col,  "Sobredimensionado?");                                  col += 1
    _hdr(ws, 1, col,  "Peligrosa?");                                          col += 1
    _hdr(ws, 1, col,  "Días de Almacenaje\nTOTALES Facturados");              col += 1
    _hdr(ws, 1, col,  "Días de Almacenaje\na Condonar por\nResponsabilidad CMSA"); col += 1
    _hdr(ws, 1, col,  "Cuántos Días le\nProceden para Cobro");                col += 1
    _hdr(ws, 1, col,  "Días de Conexiones\nTotales");                         col += 1
    _hdr(ws, 1, col,  "Días de Conexiones\na Condonar");                      col += 1
    _hdr(ws, 1, col,  "Días de Conexiones\na Pagar");                         col += 1
    _hdr(ws, 1, col,  "Monto a Condonar\npor Almacenaje");                    col += 1
    _hdr(ws, 1, col,  "Monto a Condonar\npor Conexiones");                    col += 1
    _hdr(ws, 1, col,  "Monto Maniobra\nNo Realizada");                        col += 1
    _hdr(ws, 1, col,  "Cantidad de\nReprogramaciones");                       col += 1
    _hdr(ws, 1, col,  "Admin Service\n¿Se Cobró en Factura?");                col += 1
    _hdr(ws, 1, col,  "¿Aplica Admin\nService a Condonar?");                  col += 1
    _hdr(ws, 1, col,  "Monto a Condonar\nServicio Admon\ny Control");         col += 1
    _hdr(ws, 1, col,  "Facturas\nLiquidadas");                                col += 1
    _hdr(ws, 1, col,  "A.A.");                                                col += 1
    _hdr(ws, 1, col,  "Cliente");                                             col += 1
    _hdr(ws, 1, col,  "Monto Total");                                         col += 1
    _hdr(ws, 1, col,  "N° de Cuenta\nBancaria");                              col += 1
    _hdr(ws, 1, col,  "Motivo del Rechazo\n(No Procedente)");                 col += 1
    _hdr(ws, 1, col,  "NC CLIENTE");

    total_cols = col

    ws.row_dimensions[1].height = 50

    # ── Filas de datos ─────────────────────────────────────────
    col_bi  = COL_BI["contenedor"]
    col_tab = COL_TAB["contenedor"]

    df_bi_idx  = df_bi.set_index(col_bi)
    df_tab_idx = df_tab.set_index(col_tab)

    contenedores  = df_bi[col_bi].tolist()
    monto_total_suma = 0.0

    for data_row, cont in enumerate(contenedores, start=2):
        bi  = df_bi_idx.loc[cont] if cont in df_bi_idx.index else {}
        tab = df_tab_idx.loc[cont] if cont in df_tab_idx.index else {}
        d   = desfases.get(cont, {})
        m   = montos.get(cont, {})

        def bi_val(key):
            try:
                v = bi.get(COL_BI[key], None)
                return None if pd.isna(v) else v
            except Exception:
                return None

        def tab_val(key):
            try:
                v = tab.get(COL_TAB[key], None)
                return None if pd.isna(v) else v
            except Exception:
                return None

        fecha_sol = fechas_por_contenedor.get(cont, fecha_solicitud_global)

        # Calcular días totales de intento programación (Regla 3 — días naturales)
        fp = to_date(bi_val("fecha_previo"))
        fpos = to_date(bi_val("fecha_posicion"))
        if fp and fpos:
            dias_intento = (fpos - fp).days + 1
        else:
            dias_intento = None

        # Desfases
        dp = d.get("desfase_previo", 0)
        df_f = d.get("desfase_ffcc", 0)
        dc = d.get("desfase_carretero", 0)

        paso_previo    = "SI" if dp > 0 else "NO"
        paso_ffcc      = "SI" if df_f > 0 else "NO"
        paso_carretero = "SI" if dc > 0 else "NO"

        # FFCC
        ffcc_raw = bi_val("ffcc")
        ffcc_val = "SI" if str(ffcc_raw).strip().upper() in ("YES", "SI", "SÍ", "S") else "NO"

        # Fecha autorización naviera
        fecha_auth = to_date(bi_val("fecha_auth_naviera"))

        # Comentario dinámico
        comentario = generar_comentario(cont, desfases, montos)

        # Conexiones — "No aplica" si es 0
        energia_qty = m.get("energia_qty", 0)
        dias_conex_total   = int(energia_qty) if energia_qty > 0 else 0
        dias_conex_cond    = m.get("dias_conex_condonar", 0) if energia_qty > 0 else 0
        dias_conex_cobrar  = m.get("dias_conex_cobrar",   0) if energia_qty > 0 else 0

        monto_total_cont = m.get("monto_total", 0.0)
        monto_total_suma += monto_total_cont

        # ── Escribir columnas ────────────────────────────────────
        c = 1
        _dat(ws, data_row, c, cont);                                                    c += 1
        _dat(ws, data_row, c, "");                                                      c += 1  # Manual
        _dat(ws, data_row, c, fecha_sol, fmt_date=True);                                c += 1
        _dat(ws, data_row, c, fecha_revision, fmt_date=True);                           c += 1
        _dat(ws, data_row, c, to_float(tab_val("importe")), money=True);                c += 1
        _dat(ws, data_row, c, bi_val("liner"));                                         c += 1
        _dat(ws, data_row, c, bi_val("line_op"));                                       c += 1
        _dat(ws, data_row, c, to_date(bi_val("time_in")), fmt_date=True);               c += 1  # TimeIn
        _dat(ws, data_row, c, bi_val("dia_time_in"));                                   c += 1  # Dia Time In
        _dat(ws, data_row, c, to_date(bi_val("fecha_revalidado")), fmt_date=True);      c += 1
        _dat(ws, data_row, c, bi_val("dia_revalidado"));                                c += 1
        # Previo — N/A si vacío
        fp_val = to_date(bi_val("fecha_previo"))
        _dat(ws, data_row, c, fp_val if fp_val else "N/A", fmt_date=bool(fp_val));      c += 1
        _dat(ws, data_row, c, bi_val("dia_previo") if fp_val else "N/A");               c += 1
        _dat(ws, data_row, c, paso_previo);                                             c += 1
        fp2 = to_date(bi_val("fecha_posicion"))
        _dat(ws, data_row, c, fp2 if fp2 else "N/A", fmt_date=bool(fp2));               c += 1
        _dat(ws, data_row, c, bi_val("dia_posicion") if fp2 else "N/A");                c += 1
        fc = to_date(bi_val("fecha_cancel"))
        _dat(ws, data_row, c, fc if fc else "N/A", fmt_date=bool(fc));                  c += 1
        _dat(ws, data_row, c, to_int(bi_val("no_services_order")));                     c += 1
        _dat(ws, data_row, c, dp if dp > 0 else 0);                                    c += 1
        _dat(ws, data_row, c, dias_intento if dias_intento is not None else "N/A");     c += 1
        fl = to_date(bi_val("fecha_liberacion"))
        _dat(ws, data_row, c, fl if fl else "N/A", fmt_date=bool(fl));                  c += 1
        _dat(ws, data_row, c, bi_val("dia_liberacion") if fl else "N/A");               c += 1
        _dat(ws, data_row, c, ffcc_val);                                                c += 1
        ffm = to_date(bi_val("fecha_ferromex"))
        _dat(ws, data_row, c, ffm if ffm else "N/A", fmt_date=bool(ffm));               c += 1
        _dat(ws, data_row, c, bi_val("dia_ferromex") if ffm else "N/A");                c += 1
        _dat(ws, data_row, c, paso_ffcc);                                               c += 1
        fg = to_date(bi_val("fecha_gondola"))
        _dat(ws, data_row, c, fg if fg else "N/A", fmt_date=bool(fg));                  c += 1
        _dat(ws, data_row, c, bi_val("dia_gondola") if fg else "N/A");                  c += 1
        _dat(ws, data_row, c, df_f if df_f > 0 else 0);                                c += 1
        fe = to_date(bi_val("fecha_entrega"))
        _dat(ws, data_row, c, fe if fe else "N/A", fmt_date=bool(fe));                  c += 1
        _dat(ws, data_row, c, bi_val("dia_entrega") if fe else "N/A");                  c += 1
        _dat(ws, data_row, c, fecha_auth if fecha_auth else "N/A", fmt_date=bool(fecha_auth)); c += 1
        _dat(ws, data_row, c, to_int(bi_val("no_entregas")));                           c += 1
        fce = to_date(bi_val("fecha_cmsa_entrega"))
        _dat(ws, data_row, c, fce if fce else "N/A", fmt_date=bool(fce));               c += 1
        _dat(ws, data_row, c, bi_val("dia_entrega_cmsa") if fce else "N/A");            c += 1
        _dat(ws, data_row, c, paso_carretero);                                          c += 1
        to_val = to_date(bi_val("time_out"))
        _dat(ws, data_row, c, to_val if to_val else "N/A", fmt_date=bool(to_val));      c += 1
        _dat(ws, data_row, c, bi_val("dia_time_out") if to_val else "N/A");             c += 1
        _dat(ws, data_row, c, dc if dc > 0 else 0);                                    c += 1
        _dat(ws, data_row, c, comentario);                                              c += 1
        _dat(ws, data_row, c, bi_val("tipo_contenedor"));                               c += 1
        _dat(ws, data_row, c, bi_val("is_oog"));                                        c += 1
        _dat(ws, data_row, c, bi_val("is_hazardous"));                                  c += 1
        _dat(ws, data_row, c, int(m.get("alm_qty", 0)));                                c += 1
        _dat(ws, data_row, c, int(m.get("dias_alm_condonar", 0)));                      c += 1
        _dat(ws, data_row, c, int(m.get("dias_alm_cobrar", 0)));                        c += 1
        _dat(ws, data_row, c, dias_conex_total if dias_conex_total > 0 else 0);         c += 1
        _dat(ws, data_row, c, dias_conex_cond);                                         c += 1
        _dat(ws, data_row, c, dias_conex_cobrar);                                       c += 1
        _dat(ws, data_row, c, m.get("monto_alm", 0.0), money=True);                    c += 1
        _dat(ws, data_row, c, m.get("monto_conex", 0.0), money=True);                  c += 1
        _dat(ws, data_row, c, m.get("no_show_subtotal", 0.0), money=True);              c += 1
        _dat(ws, data_row, c, m.get("no_show_qty", 0));                                 c += 1
        _dat(ws, data_row, c, m.get("admon_cobrado", "NO"));                            c += 1
        _dat(ws, data_row, c, m.get("aplica_admon", "NO"));                             c += 1
        _dat(ws, data_row, c, m.get("monto_admon", 0.0), money=True);                  c += 1
        _dat(ws, data_row, c, bi_val("no_factura"));                                    c += 1
        _dat(ws, data_row, c, bi_val("agente_aduanal"));                                c += 1
        _dat(ws, data_row, c, bi_val("cliente"));                                       c += 1
        _dat(ws, data_row, c, monto_total_cont, money=True);                            c += 1
        _dat(ws, data_row, c, tab_val("clabe"));                                        c += 1
        _dat(ws, data_row, c, "");                                                      c += 1  # Manual
        _dat(ws, data_row, c, nc_cliente);                                              c += 1

    # ── Fila de totales ────────────────────────────────────────
    fila_total = len(contenedores) + 2
    total_cell = ws.cell(row=fila_total, column=1, value="Total de Monto Procedente")
    total_cell.font = Font(bold=True, name="Arial", size=11)
    total_cell.fill = PatternFill("solid", start_color="D6E4F0")
    total_cell.alignment = _DAT_ALIGN
    total_cell.border = _BORDER

    # Buscar columna de Monto Total (col 60 en base 1)
    col_monto_total = 60
    tc = ws.cell(row=fila_total, column=col_monto_total, value=monto_total_suma)
    tc.font = Font(bold=True, name="Arial", size=11)
    tc.fill = PatternFill("solid", start_color="D6E4F0")
    tc.alignment = _DAT_ALIGN
    tc.border = _BORDER
    tc.number_format = _MONEY_FMT

    # ── Anchos de columna ──────────────────────────────────────
    anchos = {
        1: 16, 2: 14, 3: 16, 4: 14, 5: 16, 6: 10, 7: 12,
        8: 14, 9: 12, 10: 14, 11: 12, 12: 14, 13: 12, 14: 14,
        15: 14, 16: 12, 17: 12, 18: 14, 19: 14, 20: 12,
        21: 14, 22: 12, 23: 10, 24: 14, 25: 12, 26: 14,
        27: 14, 28: 12, 29: 14, 30: 14, 31: 12, 32: 18,
        33: 12, 34: 14, 35: 12, 36: 16, 37: 14, 38: 12,
        39: 14, 40: 40, 41: 14, 42: 14, 43: 12, 44: 14,
        45: 14, 46: 14, 47: 14, 48: 14, 49: 14, 50: 16,
        51: 16, 52: 16, 53: 12, 54: 14, 55: 14, 56: 16,
        57: 16, 58: 18, 59: 18, 60: 14, 61: 22, 62: 22, 63: 16,
    }
    for col_n, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(col_n)].width = ancho

    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False

    wb.save(ruta_salida)  # works with both path and BytesIO
    return ruta_salida
